#!/usr/bin/env python3
"""
jobhunt_excel.py  --  regenerates the transparent Excel tracker from the
SQLite source of truth in jobhunt_db.py.

Tier 0. No model calls. One-way: structured store -> workbook, always
regenerated in full rather than hand-patched, so the workbook can never drift
from the database silently. Two-way sync (edit the workbook, re-import,
reconcile) is a stated v2 stretch goal, not this pass -- see
`diff_workbook()` below, which validates and reports differences without
writing them back.
"""

import json
import sqlite3
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

import jobhunt_db as db

__all__ = ["generate_workbook", "default_workbook_path", "diff_workbook"]

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
HIGH_FIT_FILL = PatternFill(start_color="D1FADF", end_color="D1FADF", fill_type="solid")
STALE_FILL = PatternFill(start_color="FEE4E2", end_color="FEE4E2", fill_type="solid")


def default_workbook_path(root: Optional[Path] = None) -> Path:
    base = Path(root) if root else Path(__file__).resolve().parent
    return base / "MyData" / "jobhunt" / "JobHunt_Tracker.xlsx"


def _write_sheet(wb: Workbook, name: str, headers: List[str],
                 rows: List[Tuple[Any, ...]],
                 highlight_col: Optional[int] = None,
                 highlight_rule=None) -> Worksheet:
    ws = wb.create_sheet(title=name[:31])   # Excel's own sheet-name limit
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    for row in rows:
        ws.append(list(row))
        if highlight_col is not None and highlight_rule is not None:
            r = ws.max_row
            try:
                if highlight_rule(row[highlight_col]):
                    for cell in ws[r]:
                        cell.fill = HIGH_FIT_FILL
            except (IndexError, TypeError):
                pass
    for i, h in enumerate(headers, start=1):
        width = max(12, min(48, len(str(h)) + 4))
        ws.column_dimensions[get_column_letter(i)].width = width
    if rows:
        ws.auto_filter.ref = ws.dimensions
    return ws


def _json_or_blank(v: Any) -> str:
    if v in (None, ""):
        return ""
    if isinstance(v, (dict, list)):
        return json.dumps(v, ensure_ascii=False)
    try:
        parsed = json.loads(v)
        return json.dumps(parsed, ensure_ascii=False) if isinstance(parsed, (dict, list)) else str(v)
    except (TypeError, ValueError):
        return str(v)


def generate_workbook(conn: sqlite3.Connection, out_path: Optional[Path] = None
                      ) -> Path:
    wb = Workbook()
    wb.remove(wb.active)   # drop the default blank sheet

    opps = db.list_opportunities(conn, limit=5000)
    jobs = {j["job_id"]: j for j in db.list_jobs(conn, limit=5000)}
    companies = {c["company_id"]: c for c in db.list_companies(conn, limit=2000)}

    # 01_DASHBOARD
    snap = db.daily_snapshot(conn)
    _write_sheet(wb, "01_DASHBOARD", ["Metric", "Value"],
                [(k.replace("_", " ").title(), v) for k, v in snap.items()])

    # 02_OPPORTUNITIES
    opp_headers = ["Opportunity ID", "Company", "Role", "Route", "Status",
                  "Fit Score", "Fit Status", "Priority", "Next Action",
                  "Next Action Date", "Last Activity", "Official URL"]
    opp_rows = []
    for o in opps:
        job = jobs.get(o.get("job_id")) or {}
        company = companies.get(o.get("company_id")) or {}
        opp_rows.append((
            o["opportunity_id"], company.get("name", ""), job.get("title", ""),
            o.get("route", ""), o.get("status", ""), o.get("fit_score"),
            o.get("fit_status", ""), o.get("priority", ""),
            o.get("next_action", ""), o.get("next_action_date", ""),
            o.get("last_activity", ""), job.get("official_url", "")))
    _write_sheet(wb, "02_OPPORTUNITIES", opp_headers, opp_rows,
                highlight_col=5, highlight_rule=lambda v: isinstance(v, int) and v >= 90)

    # 03_JOBS
    job_headers = ["Job ID", "Company", "Title", "Location", "Remote", "Source Type",
                  "ATS", "Posted", "Age (days)", "Date Confidence", "Status",
                  "Official URL"]
    job_rows = [(j["job_id"], (companies.get(j.get("company_id")) or {}).get("name", ""),
               j.get("title", ""), j.get("location", ""), j.get("remote_type", ""),
               j.get("source_type", ""), j.get("ats", ""), j.get("posted_at", ""),
               j.get("age_days"), j.get("date_confidence", ""), j.get("status", ""),
               j.get("official_url", "")) for j in jobs.values()]
    _write_sheet(wb, "03_JOBS", job_headers, job_rows)

    # 04_COMPANIES
    co_headers = ["Company ID", "Name", "Domain", "Industry", "Location",
                 "Careers URL", "Research Date"]
    co_rows = [(c["company_id"], c.get("name", ""), c.get("domain", ""),
              c.get("industry", ""), c.get("location", ""),
              c.get("careers_url", ""), c.get("research_date", ""))
              for c in companies.values()]
    _write_sheet(wb, "04_COMPANIES", co_headers, co_rows)

    # 05_CONTACTS
    contacts = db.list_contacts(conn)
    ct_headers = ["Contact ID", "Name", "Company", "Title", "Email", "LinkedIn",
                 "Relationship", "Status", "Last Contact", "Next Followup"]
    ct_rows = [(c["contact_id"], c.get("name", ""),
              (companies.get(c.get("company_id")) or {}).get("name", ""),
              c.get("title", ""), c.get("email", ""), c.get("linkedin", ""),
              c.get("relationship", ""), c.get("status", ""),
              c.get("last_contact", ""), c.get("next_followup", ""))
              for c in contacts]
    _write_sheet(wb, "05_CONTACTS", ct_headers, ct_rows)

    # 06_DISCOVERY
    runs = [dict(r) for r in conn.execute(
        "SELECT * FROM discovery_runs ORDER BY started_at DESC").fetchall()]
    disc_headers = ["Run ID", "Started", "Finished", "Queries Run", "Jobs Found",
                   "Jobs Verified", "Jobs Qualified"]
    disc_rows = [(r["run_id"], r["started_at"], r["finished_at"], r["queries_run"],
                r["jobs_found"], r["jobs_verified"], r["jobs_qualified"]) for r in runs]
    _write_sheet(wb, "06_DISCOVERY", disc_headers, disc_rows)

    # 07_FIT_CHECKS
    fit_rows_raw = [dict(r) for r in conn.execute(
        "SELECT * FROM fit_checks ORDER BY created_at DESC").fetchall()]
    fit_headers = ["Fit Check ID", "Opportunity", "Score", "Category",
                  "Recommendation", "Created"]
    fit_rows = [(f["fit_check_id"], f["opportunity_id"], f["score"], f["category"],
               f["recommendation"], f["created_at"]) for f in fit_rows_raw]
    _write_sheet(wb, "07_FIT_CHECKS", fit_headers, fit_rows,
                highlight_col=2, highlight_rule=lambda v: isinstance(v, int) and v >= 90)

    # 08_RESUMES
    resumes = db.list_resume_versions(conn)
    res_headers = ["Version ID", "Job ID", "Company ID", "Base Version", "Created"]
    res_rows = [(r["version_id"], r.get("job_id", ""), r.get("company_id", ""),
               r.get("base_version_id", ""), r["created_at"]) for r in resumes]
    _write_sheet(wb, "08_RESUMES", res_headers, res_rows)

    # 09_APPLICATIONS
    app_rows = [(o["opportunity_id"],
               (companies.get(o.get("company_id")) or {}).get("name", ""),
               (jobs.get(o.get("job_id")) or {}).get("title", ""),
               o.get("application_status", ""), o.get("application_date", ""))
               for o in opps if o.get("application_status")]
    _write_sheet(wb, "09_APPLICATIONS",
                ["Opportunity ID", "Company", "Role", "Application Status",
                 "Application Date"], app_rows)

    # 10_OUTREACH
    outreach_raw = [dict(r) for r in conn.execute(
        "SELECT * FROM outreach_plans ORDER BY created_at DESC").fetchall()]
    out_headers = ["Plan ID", "Opportunity", "Channel", "Objective", "Status",
                  "Next Action", "Created"]
    out_rows = [(o["plan_id"], o["opportunity_id"], o.get("channel", ""),
               o.get("objective", ""), o.get("status", ""),
               o.get("next_action", ""), o["created_at"]) for o in outreach_raw]
    _write_sheet(wb, "10_OUTREACH", out_headers, out_rows)

    # 11_CONVERSATIONS
    conv_raw = [dict(r) for r in conn.execute(
        "SELECT * FROM conversations ORDER BY conversation_date DESC").fetchall()]
    conv_headers = ["Conversation ID", "Opportunity", "Person", "Date",
                   "Next Action", "Followup Date"]
    conv_rows = [(c["conversation_id"], c["opportunity_id"], c.get("person", ""),
                c["conversation_date"], c.get("next_action", ""),
                c.get("followup_date", "")) for c in conv_raw]
    _write_sheet(wb, "11_CONVERSATIONS", conv_headers, conv_rows)

    # 12_TASKS
    tasks = [dict(r) for r in conn.execute("SELECT * FROM tasks ORDER BY due_date").fetchall()]
    task_headers = ["Task ID", "Opportunity", "Title", "Due Date", "Status"]
    task_rows = [(t["task_id"], t.get("opportunity_id", ""), t["title"],
                t.get("due_date", ""), t["status"]) for t in tasks]
    _write_sheet(wb, "12_TASKS", task_headers, task_rows)

    # 13_FOLLOWUPS
    followups = [dict(r) for r in conn.execute("SELECT * FROM followups ORDER BY due_date").fetchall()]
    fup_headers = ["Followup ID", "Opportunity", "Due Date", "Reason", "Status"]
    fup_rows = [(f["followup_id"], f["opportunity_id"], f["due_date"],
               f.get("reason", ""), f["status"]) for f in followups]
    _write_sheet(wb, "13_FOLLOWUPS", fup_headers, fup_rows)

    # 14_STATUS_HISTORY
    hist = [dict(r) for r in conn.execute(
        "SELECT * FROM status_history ORDER BY at DESC LIMIT 2000").fetchall()]
    hist_headers = ["Opportunity", "From", "To", "Note", "At"]
    hist_rows = [(h["opportunity_id"], h.get("from_status", ""), h["to_status"],
                h.get("note", ""), h["at"]) for h in hist]
    _write_sheet(wb, "14_STATUS_HISTORY", hist_headers, hist_rows)

    # 15_DAILY_UPDATES -- one row per generation, appended by jobhunt_daily.py
    daily_headers = ["Date", "New Discoveries", "Qualified", "High Value (90+)",
                     "Overdue Tasks", "Followups Due"]
    daily_rows = [(snap["date"], snap["new_discoveries_today"], snap["qualified_jobs"],
                 snap["high_value_90plus"], snap["overdue_tasks"], snap["followups_due"])]
    _write_sheet(wb, "15_DAILY_UPDATES", daily_headers, daily_rows)

    # 16_OUTCOMES
    outcome_rows = [(o["opportunity_id"],
                   (companies.get(o.get("company_id")) or {}).get("name", ""),
                   o.get("outcome", ""), o.get("status", ""))
                   for o in opps if o.get("status") in
                   ("OFFER", "REJECTED", "WITHDRAWN", "CLOSED")]
    _write_sheet(wb, "16_OUTCOMES",
                ["Opportunity ID", "Company", "Outcome", "Final Status"], outcome_rows)

    # 17_ROLE_PERMUTATIONS
    perms = db.get_role_permutations(conn)
    perm_headers = ["Sheet", "Canonical Role", "Designation", "Alt Designation",
                   "Seniority", "Function", "Include/Exclude", "Priority"]
    perm_rows = [(p.get("sheet", ""), p.get("canonical_role", ""),
                p.get("designation", ""), p.get("alternative_designation", ""),
                p.get("seniority", ""), p.get("function", ""),
                p.get("include_exclude", ""), p.get("search_priority", ""))
                for p in perms]
    _write_sheet(wb, "17_ROLE_PERMUTATIONS", perm_headers, perm_rows)

    # 18_SEARCH_QUERIES
    queries = [dict(r) for r in conn.execute(
        "SELECT * FROM search_queries ORDER BY at DESC LIMIT 2000").fetchall()]
    q_headers = ["Query", "Provider", "Result State", "Result Count", "At"]
    q_rows = [(q["query_text"], q["provider"], q["result_state"],
              q["result_count"], q["at"]) for q in queries]
    _write_sheet(wb, "18_SEARCH_QUERIES", q_headers, q_rows)

    # 19_SETTINGS
    _write_sheet(wb, "19_SETTINGS", ["Setting", "Value"], [
        ("MAX_JOB_AGE_DAYS", 7),
        ("FIT_QUALIFY_THRESHOLD", 90),
        ("Generated by", "jobhunt_excel.py (one-way, structured store -> workbook)"),
    ])

    # 20_AUDIT_LOG
    audit_rows_raw = [dict(r) for r in conn.execute(
        "SELECT * FROM audit_log ORDER BY at DESC LIMIT 5000").fetchall()]
    audit_headers = ["Entity Type", "Entity ID", "Action", "Detail", "At"]
    audit_rows = [(a["entity_type"], a.get("entity_id", ""), a["action"],
                 a.get("detail", ""), a["at"]) for a in audit_rows_raw]
    _write_sheet(wb, "20_AUDIT_LOG", audit_headers, audit_rows)

    out = Path(out_path) if out_path else default_workbook_path()
    out.parent.mkdir(parents=True, exist_ok=True)
    tmp = out.with_suffix(".tmp.xlsx")
    wb.save(str(tmp))
    tmp.replace(out)   # atomic-ish swap, never leaves a half-written workbook
    return out


def diff_workbook(conn: sqlite3.Connection, path: Path) -> Dict[str, Any]:
    """Report-only import (INGESTION.md: report by default, act only when
    told to). Reads a hand-edited workbook's 02_OPPORTUNITIES sheet and
    reports what differs from the database, without writing anything back.
    Two-way reconciliation is a v2 stretch goal -- this only ever reports.
    """
    from openpyxl import load_workbook

    path = Path(path)
    if not path.is_file():
        return {"ok": False, "error": "workbook not found: %s" % path}

    wb = load_workbook(str(path), data_only=True, read_only=True)
    if "02_OPPORTUNITIES" not in wb.sheetnames:
        return {"ok": False, "error": "02_OPPORTUNITIES sheet not found"}

    ws = wb["02_OPPORTUNITIES"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    known = {o["opportunity_id"]: o for o in db.list_opportunities(conn, limit=5000)}

    differences: List[Dict[str, Any]] = []
    seen_ids = set()
    for row in rows:
        if not row or not row[0]:
            continue
        oid = str(row[0])
        seen_ids.add(oid)
        current = known.get(oid)
        if current is None:
            differences.append({"opportunity_id": oid, "issue": "not in database"})
            continue
        sheet_status = row[4] if len(row) > 4 else None
        if sheet_status and sheet_status != current.get("status"):
            differences.append({
                "opportunity_id": oid, "issue": "status differs",
                "workbook": sheet_status, "database": current.get("status")})

    missing_from_sheet = [oid for oid in known if oid not in seen_ids]
    return {"ok": True, "differences": differences,
           "missing_from_workbook": missing_from_sheet,
           "note": "report only, nothing was written back to the database"}
