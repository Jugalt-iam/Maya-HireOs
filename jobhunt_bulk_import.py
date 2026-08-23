#!/usr/bin/env python3
"""
jobhunt_bulk_import.py  --  bulk import of externally produced "Results"
spreadsheets (Company Name / Designation / Tier-Match / Location /
Discovery Source / Company URL / Notes) into the structured Job Hunt
database.

Tier 0, no model. This is not Route 1 (Discovery) finding a job on its own
-- it is the user handing the system a set of roles they already found and
are vouching for personally, compiled into a spreadsheet instead of pasted
one at a time. Filed under route INBOUND: a human telling the system about
an opportunity, the same shape as a recruiter message, just batched.

What this does NOT do:
  - Invent a fit score. The sheet's own "Tier / Match" column is an earlier,
    informal read, not this system's scored evidence classification
    (jobhunt_fit.py) -- it is preserved verbatim in opportunities.notes,
    labeled as self-reported, and fit_score stays NULL until a real Fit
    Check runs against the actual resume.
  - Trust a link just because a human trusts the role. Every candidate URL
    still goes through jobhunt_verify.classify_url(), the same function
    every other intake route uses. official_url is only ever set to a link
    that mechanically resolves to an OFFICIAL_ATS or COMPANY_SITE source.
    An unverified row is not rejected -- it is created as DISCOVERED, same
    as jobhunt_from_portal does for a link that fails verification, so it
    stays visible and ready for a real posting link to move it forward.

Reads only sheets whose header row has both a company-name and a
designation/title column (tolerant of minor header-text differences --
the two workbooks that prompted this spell the discovery-source column
slightly differently). Sheets that don't match (an "Excluded ..."
kill-list sheet, a "Method and ..." narrative sheet) are skipped, not
guessed at.

Run directly against the real database:
    python jobhunt_bulk_import.py "Round_3.xlsx" "Volume_Build.xlsx"
"""

import re
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from urllib.parse import urlparse

from openpyxl import load_workbook

import jobhunt_db as db
import jobhunt_verify as verify

__all__ = ["import_workbook", "import_workbooks"]

_COLUMN_ALIASES = {
    "company name": "company",
    "company": "company",
    "designation": "title",
    "role": "title",
    "title": "title",
    "tier / match": "tier",
    "tier/match": "tier",
    "tier": "tier",
    "location": "location",
    "company url": "company_url",
    "companyurl": "company_url",
}


def _normalized_header(cell: Any) -> str:
    return re.sub(r"\s+", " ", str(cell or "")).strip().lower()


def _map_headers(row: Tuple[Any, ...]) -> Optional[Dict[str, int]]:
    """Maps this row's cells to field names by header text. Returns None if
    the row isn't a Results-shaped header.

    An "Excluded ..." kill-list sheet also has "Company" and "Designation"
    columns in both source workbooks -- company/title alone is not enough
    to tell it apart from a real Results sheet, and importing a kill-list
    sheet as if it were open opportunities would be a real, serious
    correctness bug (it would recreate exactly the roles the user already
    marked "already applied" or "already rejected" as if they were new).
    Two more signals make the distinction reliable: a Results sheet always
    also carries a "Discovery Source" column that a kill-list sheet never
    has, and a kill-list sheet always carries a "Reason excluded" column
    that a Results sheet never has. Both are checked, not just one.
    """
    mapping: Dict[str, int] = {}
    for i, cell in enumerate(row):
        h = _normalized_header(cell)
        if not h:
            continue
        if h.startswith("reason"):
            return None   # a kill-list/excluded sheet, never a Results sheet
        if h in _COLUMN_ALIASES:
            mapping[_COLUMN_ALIASES[h]] = i
        elif h.startswith("discovery source"):
            mapping["discovery_source"] = i
        elif h == "notes":
            mapping["notes"] = i
    if not all(k in mapping for k in ("company", "title", "discovery_source")):
        return None
    return mapping


def _row_value(row: Tuple[Any, ...], mapping: Dict[str, int], key: str) -> str:
    idx = mapping.get(key)
    if idx is None or idx >= len(row):
        return ""
    v = row[idx]
    return "" if v is None else str(v).strip()


def _domain_from_url(url: str) -> str:
    try:
        u = url if "://" in url else "https://" + url
        host = (urlparse(u).hostname or "").lower()
    except ValueError:
        return ""
    return host[4:] if host.startswith("www.") else host


def import_workbook(conn, path: Path, source_label: str = "") -> Dict[str, Any]:
    """Imports every Results-shaped sheet in one workbook. Returns a report
    -- counts plus one line per row, never silent about a row that was
    skipped, deduplicated, or created unverified."""
    path = Path(path)
    label = source_label or path.name
    report: Dict[str, Any] = {
        "ok": True, "file": path.name, "sheets_read": [], "sheets_skipped": [],
        "created": 0, "deduplicated": 0, "verified": 0, "unverified": 0,
        "rows": [],
    }
    try:
        wb = load_workbook(str(path), read_only=True, data_only=True)
    except Exception as exc:
        return {"ok": False, "file": path.name, "error": str(exc)}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            continue
        mapping = _map_headers(header)
        if mapping is None:
            report["sheets_skipped"].append(sheet_name)
            continue
        report["sheets_read"].append(sheet_name)

        for row in rows_iter:
            if row is None or not any(row):
                continue
            company = _row_value(row, mapping, "company")
            title = _row_value(row, mapping, "title")
            if not (company and title):
                continue
            result = _import_row(
                conn, company=company, title=title,
                location=_row_value(row, mapping, "location"),
                tier=_row_value(row, mapping, "tier"),
                discovery_source=_row_value(row, mapping, "discovery_source"),
                company_url=_row_value(row, mapping, "company_url"),
                notes=_row_value(row, mapping, "notes"),
                source_label=label)
            report["rows"].append(result)
            if result["outcome"] == "created":
                report["created"] += 1
                report["verified" if result["verified"] else "unverified"] += 1
            elif result["outcome"] == "deduplicated":
                report["deduplicated"] += 1
    return report


def _import_row(conn, *, company: str, title: str, location: str, tier: str,
                discovery_source: str, company_url: str, notes: str,
                source_label: str) -> Dict[str, Any]:
    domain = _domain_from_url(company_url) if company_url else None
    classification = verify.classify_url(discovery_source or "", domain)
    verified = bool(discovery_source) and verify.is_official(classification)
    # company_url is the company's own site, never treated as if it were a
    # specific posting -- only a link that verifies as OFFICIAL_ATS or the
    # company's own careers-shaped path becomes official_url.
    official_url = discovery_source if verified else None

    signature = db.dedup_signature(company, title, location, official_url or "")
    existing = db.find_opportunity_by_signature(conn, signature)
    if existing:
        return {"company": company, "title": title, "outcome": "deduplicated",
               "opportunity_id": existing["opportunity_id"], "verified": verified}

    note_parts = []
    if tier:
        note_parts.append("Self-reported tier (not a computed Fit Score): %s" % tier)
    if notes:
        note_parts.append(notes)
    if discovery_source and not verified:
        note_parts.append("Discovery link (not verified as a direct posting -- "
                          "%s): %s" % (classification.get("source_type", "UNKNOWN"),
                                      discovery_source))
    note_parts.append("Imported from %s." % source_label)

    company_fields = {"domain": domain} if domain else {}
    company_id = db.upsert_company(conn, company, **company_fields)
    job_id = db.create_job(
        conn, company_id, title, location=location,
        official_url=official_url, source_url=discovery_source or company_url,
        source_type=classification.get("source_type"),
        ats=classification.get("ats"), date_confidence="UNKNOWN",
        status="VERIFIED" if verified else "DISCOVERED")
    job = db.get_job(conn, job_id)
    opportunity_id = db.create_opportunity(
        conn, company_id, "INBOUND", job_id=job_id,
        dedup_signature=job["dedup_signature"],
        status="VERIFIED" if verified else "DISCOVERED",
        notes=" | ".join(note_parts))
    return {"company": company, "title": title, "outcome": "created",
           "opportunity_id": opportunity_id, "job_id": job_id, "verified": verified}


def import_workbooks(conn, paths: List[Path]) -> Dict[str, Any]:
    results = [import_workbook(conn, p) for p in paths]
    return {
        "ok": all(r.get("ok") for r in results),
        "files": results,
        "created": sum(r.get("created", 0) for r in results),
        "deduplicated": sum(r.get("deduplicated", 0) for r in results),
        "verified": sum(r.get("verified", 0) for r in results),
        "unverified": sum(r.get("unverified", 0) for r in results),
    }


def _cli() -> int:
    if len(sys.argv) < 2:
        print("usage: python jobhunt_bulk_import.py <workbook.xlsx> [more.xlsx ...]")
        return 2
    conn = db.connect()
    db.init_schema(conn)
    result = import_workbooks(conn, [Path(p) for p in sys.argv[1:]])
    for file_report in result["files"]:
        if not file_report.get("ok"):
            print("FAILED  %s -- %s" % (file_report.get("file"), file_report.get("error")))
            continue
        print("\n%s" % file_report["file"])
        print("  sheets read: %s" % file_report["sheets_read"])
        print("  sheets skipped (not Results-shaped): %s" % file_report["sheets_skipped"])
        for r in file_report["rows"]:
            tag = "verified  " if r.get("verified") else "unverified"
            marker = {"created": "NEW ", "deduplicated": "DUP "}.get(r["outcome"], "?   ")
            print("  %s%s %-28s %-45s -> %s" % (
                marker, tag, r["company"][:28], r["title"][:45],
                r.get("opportunity_id", "")))
    print("\nTotal: %d created (%d verified, %d unverified), %d already existed."
         % (result["created"], result["verified"], result["unverified"],
            result["deduplicated"]))
    return 0 if result["ok"] else 1


if __name__ == "__main__":
    raise SystemExit(_cli())
